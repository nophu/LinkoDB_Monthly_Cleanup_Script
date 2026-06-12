import json, os, re
from datetime import date
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import REPORT_CONFIG

# ── style ─────────────────────────────────────────────────────────────────────
GREEN  = "92D050"
YELLOW = "FFFF00"
ONEOFF = "BDD7EE"
GREY   = "D9D9D9"
RED    = "FF0000"
BLUE   = "0070C0"
FONT   = "Aptos Narrow"

# friendly display names for the new "Field" column
FIELD_DISPLAY = {
    "Extractor ID":        "Extractor ID",
    "Extractor Type":      "Extractor Type",
    "Trap Size and Units": "Trap Size & Units",
    "Cleaning Frequency":  "Cleaning Frequency",
    "ReceivingPlant":      "Receiving Plant",
    "ClassCode":           "Class Code",
    "SecondClass":         "Second Class",
    "TrunkLine":           "Trunk Line",
    "MapCategory":         "Map Category",
    "EventTypeAbbrv":      "Event Type",
}

def _f(size=11, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def _fill(c): return PatternFill("solid", fgColor=c)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, value, bold=False, color="000000", fill=None,
          halign="left", italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _f(11, bold, color, italic)
    c.alignment = Alignment(horizontal=halign, vertical="center",
                            indent=1 if halign == "left" else 0, wrap_text=False)
    if fill: c.fill = _fill(fill)
    c.border = _border()
    return c


# columns: A=label  B=facility  C=permit  D=field  E=current  F=changed to
COL_LABEL, COL_FAC, COL_PERMIT, COL_FIELD, COL_CUR, COL_CHG, COL_REASON = 1, 2, 3, 4, 5, 6, 7


def _suggestion(change):
    if change["status"] == "fixed":
        return change["cleaned_value"]
    note = change.get("note", "")
    m = re.search(r"should it be changed to '([^']+)'", note)
    if m: return m.group(1)
    m = re.search(r"suggested: '([^']+)'", note)
    if m: return m.group(1)
    if "leave this field blank" in note or "should be deleted" in note:
        return ""   # actual blank cell instead of the word "Blank"
    return "Needs Manual Review"


# one issue per row, deduplicated by (facility, field, original value)
def _issue_rows(changes, field_order):
    seen = set()
    rows = []
    for chg in changes:
        key = (chg["facility"], chg["permit_no"], chg["field"], chg["original"])
        if key in seen: continue
        seen.add(key)
        rows.append({
            "facility": chg["facility"],
            "permit":   chg["permit_no"],
            "field":    chg["field"],
            "current":  chg["original"],
            "changed":  _suggestion(chg),
            "status":   chg["status"],
            "note":     chg.get("note", ""),
        })
    # sort by facility, then rubric field order
    def sk(r):
        try: fi = field_order.index(r["field"])
        except ValueError: fi = 99
        return (r["facility"].lower(), fi)
    rows.sort(key=sk)
    return rows


# A "one-off" (blue) is either:
#  (a) a value that recurs across the data but never appears in the rubric
#      — e.g. "Quasar" in MapCategory (appears many times, never valid); or
#  (b) a forced exception we've been told to always treat as a one-off
#      — currently the HSW / Septic disposal stations.
# Everything else that's flagged stays yellow (review).
ONEOFF_EXCEPTIONS = ("hsw", "septic")

def _is_oneoff(r, recurring):
    orig = str(r["current"]).lower()
    if any(p in orig for p in ONEOFF_EXCEPTIONS):
        return True
    return ("is not a valid value for" in r.get("note", "")
            and (r["field"], r["current"]) in recurring)


# Plain-language justification for every change, derived from the validator's note
# plus the one-off classification. Shown in the report's "Reason" column.
def _reason(r, is_oneoff):
    note = r.get("note", "")
    if r["status"] == "fixed":
        if "fixed casing" in note:          return "Casing corrected to the rubric's accepted spelling."
        if "added EX prefix" in note:       return "Added the required 'EX' prefix; value was otherwise valid."
        if "migrated old" in note:
            m = re.search(r"migrated old (.+?) ID", note)
            t = m.group(1) if m else "type"
            return f"Old ID mapped to the new series — description identifies a {t}."
        if "mapped '" in note:              return "Program equivalent of a valid rubric value."
        if "corrected unit by rule" in note:return "Unit set by rubric rule (size ≤99 → gpm, ≥100 → gal)."
        return "Auto-corrected to a valid rubric value."
    # flagged
    if any(p in str(r["current"]).lower() for p in ONEOFF_EXCEPTIONS):
        return "Disposal station — treated as a one-off per program guidance."
    if is_oneoff:
        return "Recurs in the data but never appears in the rubric — one-off to decide on."
    if "per the old rubric" in note:
        return "The old code's number identifies its type per the old rubric — confirm the suggested new ID."
    if "material isn't stated" in note:
        return "Old grease interceptor, but the material isn't stated — choose EX100 (concrete) or EX120 (other)."
    if "no type stated" in note or "old-scheme ID" in note:
        return "Old-scheme ID with no type stated — assign a new EX number manually."
    if "is not in any valid range" in note:
        return "ID number is outside every valid EX range."
    if "is not a standard extractor ID" in note:
        return "Not a standard extractor-ID format."
    if "should be deleted" in note or "leave this field blank" in note:
        return "Rubric requires this field to be blank."
    if "expected format" in note or "not a valid numeric" in note:
        return "Value doesn't match the expected format."
    if "is not a valid value for" in note:
        return "Not a valid rubric value and doesn't recur — manual review."
    return "Needs manual review."


def _write_block(ws, row, cfg, changes, write_headers, recurring=frozenset()):
    fields = cfg["fields"]

    # Report Name row (+ column headers on the first block)
    _cell(ws, row, COL_LABEL, "Report Name:", bold=True, color=BLUE)
    _cell(ws, row, COL_FAC,   cfg["name"], bold=True)
    if write_headers:
        _cell(ws, row, COL_PERMIT, "Permit No",  bold=True, halign="center", fill=GREY)
        _cell(ws, row, COL_FIELD,  "Field",      bold=True, halign="center", fill=GREY)
        _cell(ws, row, COL_CUR,    "Current",    bold=True, halign="center", fill=GREY)
        _cell(ws, row, COL_CHG,    "Changed To", bold=True, halign="center", fill=GREY)
        _cell(ws, row, COL_REASON, "Reason",     bold=True, halign="center", fill=GREY)
    ws.row_dimensions[row].height = 18
    row += 1

    # Fields Checked row
    _cell(ws, row, COL_LABEL, "Fields Checked:", bold=True, color=BLUE)
    friendly = [FIELD_DISPLAY.get(f, f) for f in fields]
    _cell(ws, row, COL_FAC, ",  ".join(friendly))
    ws.row_dimensions[row].height = 16
    row += 1

    # Issues Found rows
    rows = _issue_rows(changes, fields)
    if not rows:
        _cell(ws, row, COL_LABEL, "Issues Found:", bold=True, color=BLUE)
        _cell(ws, row, COL_FAC, "None")
        ws.row_dimensions[row].height = 16
        return row + 1

    for i, r in enumerate(rows):
        is_one = _is_oneoff(r, recurring)
        if r["status"] == "fixed":  fill = GREEN
        elif is_one:                fill = ONEOFF   # recurring value missing from rubric
        else:                       fill = YELLOW   # can't guess, not a one-off
        _cell(ws, row, COL_LABEL, "Issues Found:" if i == 0 else "",
              bold=(i == 0), color=BLUE)
        _cell(ws, row, COL_FAC,    r["facility"], fill=fill)
        _cell(ws, row, COL_PERMIT, r["permit"],   fill=fill, halign="center")
        _cell(ws, row, COL_FIELD,  FIELD_DISPLAY.get(r["field"], r["field"]),
              fill=fill, bold=True)
        _cell(ws, row, COL_CUR,    r["current"], fill=fill, color=RED)
        nmr = (r["changed"] == "May Need Manual Review")
        _cell(ws, row, COL_CHG,    r["changed"], fill=fill,
              color=(RED if r["status"] == "flagged" else "000000"),
              italic=nmr)
        _cell(ws, row, COL_REASON, _reason(r, is_one), fill=fill, italic=True)
        ws.row_dimensions[row].height = 15
        row += 1

    return row


def build_report(changes_path="output/all_changes.json",
                 output_path="output/Monthly_Quality_Check_Report.xlsx",
                 only_reports=None):
    with open(changes_path) as f:
        all_changes = json.load(f)

    # A value is a one-off only if it recurs (appears 2+ times) AND never appears
    # in the rubric (flagged as "is not a valid value for ..."). Count globally so a
    # value like "Quasar" that recurs across facilities is recognised as a one-off.
    _unmatched = Counter()
    for c in all_changes:
        if "is not a valid value for" in c.get("note", ""):
            _unmatched[(c["field"], c["original"])] += 1
    recurring = {k for k, n in _unmatched.items() if n >= 2}

    by_file = defaultdict(list)
    for c in all_changes:
        by_file[c["source_file"]].append(c)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Quality Check {date.today().strftime('%b %Y')}"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"Linko DB Monthly Quality Check  —  {date.today().strftime('%B %d, %Y')}"
    t.font = _f(13, bold=True)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # ── color key / legend ──────────────────────────────────────────────────
    row = 2
    _cell(ws, row, COL_LABEL, "Color Key:", bold=True)
    row += 1

    _cell(ws, row, COL_FAC, "Green", bold=True, fill=GREEN, halign="center")
    ws.merge_cells(start_row=row, start_column=COL_PERMIT, end_row=row, end_column=COL_REASON)
    _cell(ws, row, COL_PERMIT,
          "Auto-fixed — the tool applied a clear correction (casing, EX prefix, trap unit). Apply it as-is in Linko.")
    ws.row_dimensions[row].height = 15
    row += 1

    _cell(ws, row, COL_FAC, "Yellow", bold=True, fill=YELLOW, halign="center")
    ws.merge_cells(start_row=row, start_column=COL_PERMIT, end_row=row, end_column=COL_REASON)
    _cell(ws, row, COL_PERMIT,
          "Needs review — no auto-fix and not a one-off (e.g. old ID needing a new number).")
    ws.row_dimensions[row].height = 15
    row += 1

    _cell(ws, row, COL_FAC, "Blue", bold=True, fill=ONEOFF, halign="center")
    ws.merge_cells(start_row=row, start_column=COL_PERMIT, end_row=row, end_column=COL_REASON)
    _cell(ws, row, COL_PERMIT,
          "One-off — recurs in the data but never in the rubric (e.g. Quasar), plus HSW/Septic stations.")
    ws.row_dimensions[row].height = 15
    row += 2  # blank separator before first report

    items = list(REPORT_CONFIG.items())
    if only_reports:
        items = [(k, c) for k, c in items if k in only_reports]

    for idx, (key, cfg) in enumerate(items):
        row = _write_block(ws, row, cfg, by_file.get(key, []), write_headers=(idx == 0))
        row += 1  # blank separator

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 26
    ws.column_dimensions["G"].width = 52
    ws.freeze_panes = "B2"

    wb.save(output_path)
    print(f"Saved: {output_path}")
    return output_path


if __name__ == "__main__":
    os.makedirs("output", exist_ok=True)
    build_report()